from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Literal

import openpyxl
import pillow_avif  # noqa: F401 - side-effect import, registers the AVIF codec with Pillow
import xmltodict
import yaml
from PIL import Image
from fastapi import APIRouter, UploadFile, File, Form
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/convert", tags=["Convert"])


# ---------------------------------------------------------
# Resource budgets
# ---------------------------------------------------------
MEDIA_MAX_FILE_SIZE = int(os.getenv("CONVERT_MEDIA_MAX_FILE_SIZE", str(12 * 1024 * 1024)))
MEDIA_MAX_TOTAL_INPUT_SIZE = int(os.getenv("CONVERT_MEDIA_MAX_TOTAL_INPUT_SIZE", str(40 * 1024 * 1024)))
MEDIA_MAX_TOTAL_OUTPUT_SIZE = int(os.getenv("CONVERT_MEDIA_MAX_TOTAL_OUTPUT_SIZE", str(64 * 1024 * 1024)))
DATA_MAX_FILE_SIZE = int(os.getenv("CONVERT_DATA_MAX_FILE_SIZE", str(16 * 1024 * 1024)))
DOC_MAX_FILE_SIZE = int(os.getenv("CONVERT_DOC_MAX_FILE_SIZE", str(50 * 1024 * 1024)))
DOC_MAX_OUTPUT_SIZE = int(os.getenv("CONVERT_DOC_MAX_OUTPUT_SIZE", str(100 * 1024 * 1024)))
CONVERSION_MAX_CONCURRENCY = max(1, int(os.getenv("CONVERT_MAX_CONCURRENCY", "2")))
LIBREOFFICE_TIMEOUT_SECONDS = max(1, int(os.getenv("CONVERT_LIBREOFFICE_TIMEOUT_SECONDS", "120")))
IMAGE_MAX_PIXELS = max(1, int(os.getenv("CONVERT_IMAGE_MAX_PIXELS", "40000000")))

Image.MAX_IMAGE_PIXELS = IMAGE_MAX_PIXELS
_conversion_slots = asyncio.Semaphore(CONVERSION_MAX_CONCURRENCY)


# ---------------------------------------------------------
# Unified API error helper
# ---------------------------------------------------------
def api_error(code: str, message: str, status: int = 400, field: str | None = None):
    detail = {"code": code, "message": message}
    if field:
        detail["field"] = field
    from fastapi import HTTPException
    raise HTTPException(status_code=status, detail=detail)


# ---------------------------------------------------------
# Helpers
# ---------------------------------------------------------
IMAGE_INPUT_EXTS = {"png", "jpg", "jpeg", "webp", "gif", "bmp", "tif", "tiff", "avif"}
IMAGE_TARGET_EXTS = {"png", "jpg", "jpeg", "webp", "gif", "bmp", "tiff", "avif"}

DATA_INPUT_EXTS = {"csv", "tsv", "json", "xml", "xlsx", "yaml", "yml"}
DATA_TARGET_EXTS = {"csv", "tsv", "json", "xml", "xlsx", "yaml", "yml"}

DOC_INPUT_EXTS = {"docx", "pdf"}
DOC_TARGET_EXTS = {"docx", "pdf"}


def safe_ext(filename: str) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    return ext


def base_name(filename: str) -> str:
    return Path(filename).stem


async def read_upload_limited(upload: UploadFile, max_size: int, field: str = "file") -> bytes:
    """Read an upload in bounded chunks and fail before arbitrary request bodies fill RAM."""
    data = bytearray()
    while True:
        chunk = await upload.read(1024 * 1024)
        if not chunk:
            break
        if len(data) + len(chunk) > max_size:
            api_error(
                "FILE_TOO_LARGE",
                f"Максимальный размер файла: {max_size} байт.",
                status=413,
                field=field,
            )
        data.extend(chunk)
    return bytes(data)


async def close_upload(upload: UploadFile):
    try:
        await upload.close()
    except Exception:
        pass


def guess_mime_for_ext(ext: str) -> str:
    ext = ext.lower()
    if ext == "pdf":
        return "application/pdf"
    if ext == "docx":
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if ext == "xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if ext == "json":
        return "application/json"
    if ext == "xml":
        return "application/xml"
    if ext == "csv":
        return "text/csv"
    if ext == "tsv":
        return "text/tab-separated-values"
    if ext in {"yaml", "yml"}:
        return "application/yaml"
    if ext in {"jpg", "jpeg"}:
        return "image/jpeg"
    if ext == "png":
        return "image/png"
    if ext == "webp":
        return "image/webp"
    if ext == "gif":
        return "image/gif"
    if ext == "bmp":
        return "image/bmp"
    if ext in {"tif", "tiff"}:
        return "image/tiff"
    if ext == "avif":
        return "image/avif"
    return "application/octet-stream"


def zip_bytes(outputs: list[tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in outputs:
            z.writestr(name, data)
    buf.seek(0)
    return buf.read()


def streaming_download(data: bytes, filename: str, mime: str):
    return StreamingResponse(
        io.BytesIO(data),
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------
# Image conversion
# ---------------------------------------------------------
def convert_image_bytes(input_bytes: bytes, src_ext: str, target_ext: str) -> bytes:
    src_ext = src_ext.lower()
    target_ext = target_ext.lower()

    with Image.open(io.BytesIO(input_bytes)) as img:
        # Normalize mode for formats
        if target_ext in {"jpg", "jpeg"}:
            # JPEG doesn't support alpha; flatten if needed
            if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
                bg = Image.new("RGB", img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
                img = bg
            else:
                img = img.convert("RGB")

        out = io.BytesIO()
        if target_ext in {"jpg", "jpeg"}:
            save_format = "JPEG"
        elif target_ext == "tif":
            save_format = "TIFF"
        else:
            save_format = target_ext.upper()

        # Some sane defaults
        save_kwargs = {}
        if save_format == "JPEG":
            save_kwargs.update({"quality": 92, "optimize": True})
        if save_format == "WEBP":
            save_kwargs.update({"quality": 90, "method": 6})

        img.save(out, format=save_format, **save_kwargs)
        return out.getvalue()


# ---------------------------------------------------------
# Data conversions
#
# Every format is parsed into one canonical Python object (dict/list/scalar,
# same shape json.loads would give you) and every target is serialized from
# that same object. This is what lets any-input-to-any-target work (7 formats
# would otherwise be 42 bespoke bytes-to-bytes functions) and it's also what
# makes every conversion consistently error as a clean api_error instead of
# an unhandled 500 - all parsing and serializing goes through these functions.
# ---------------------------------------------------------
def _unwrap_for_table(obj):
    # XML round-trips wrap arrays as {"root": {"item": [...]}} (see
    # serialize_xml_bytes) since XML has no bare-array shape. Unwrap that
    # nesting so XML -> CSV/TSV/XLSX still finds the underlying row list
    # instead of rejecting it as "not a list".
    while isinstance(obj, dict) and len(obj) == 1:
        obj = next(iter(obj.values()))
    return obj


def _row_dicts_to_table(rows) -> tuple[list[str], list[list]]:
    rows = _unwrap_for_table(rows)
    if not isinstance(rows, list):
        api_error(
            "INVALID_DATA",
            "Табличный формат (CSV/TSV/XLSX) требует список объектов (JSON-массив) на входе.",
            status=422,
        )

    headers: list[str] = []
    seen = set()
    for row in rows:
        if isinstance(row, dict):
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    headers.append(k)

    table = [[row.get(h) if isinstance(row, dict) else None for h in headers] for row in rows]
    return headers, table


def parse_delimited_bytes(raw: bytes, delimiter: str) -> list:
    text = raw.decode("utf-8-sig", errors="replace")
    try:
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        return list(reader)
    except csv.Error as e:
        api_error("INVALID_CSV", f"Не удалось разобрать табличные данные: {e}", status=422)


def serialize_delimited_bytes(obj, delimiter: str) -> bytes:
    headers, table = _row_dicts_to_table(obj)
    out = io.StringIO()
    writer = csv.writer(out, delimiter=delimiter)
    writer.writerow(headers)
    writer.writerows(table)
    return out.getvalue().encode("utf-8")


def parse_json_bytes(raw: bytes):
    try:
        return json.loads(raw.decode("utf-8", errors="replace"))
    except json.JSONDecodeError as e:
        api_error("INVALID_JSON", f"Некорректный JSON: {e}", status=422)


def serialize_json_bytes(obj) -> bytes:
    return json.dumps(obj, ensure_ascii=False, indent=2).encode("utf-8")


def parse_xml_bytes(raw: bytes):
    text = raw.decode("utf-8", errors="replace")
    try:
        return xmltodict.parse(text)
    except Exception as e:
        api_error("INVALID_XML", f"Некорректный XML: {e}", status=422)


def serialize_xml_bytes(obj) -> bytes:
    # xmltodict needs exactly one root tag name mapped to its content. A dict
    # with one key is used as-is (its value becomes the root's content), so a
    # round-tripped XML document keeps its original root tag name. Anything
    # else (a bare list/scalar, or a dict with != 1 key, e.g. from CSV/JSON/
    # YAML) is wrapped under a synthetic "root" tag.
    if isinstance(obj, dict) and len(obj) == 1:
        root_obj = obj
    else:
        root_obj = {"root": obj}

    # A list can't sit directly as a root's value - xmltodict has no element
    # name to give each item - so nest it one level deeper under "item". This
    # is the fix for the crash on any array-shaped JSON (the single most
    # common JSON shape, including this endpoint's own CSV/XLSX/TSV -> JSON
    # output): xmltodict.unparse({"root": [...]}) used to raise
    # "document with multiple roots".
    root_key, root_val = next(iter(root_obj.items()))
    if isinstance(root_val, list):
        root_obj = {root_key: {"item": root_val}}

    try:
        xml = xmltodict.unparse(root_obj, pretty=True)
    except Exception as e:
        api_error("XML_EXPORT_FAILED", f"Не удалось сериализовать в XML: {e}", status=422)
    return xml.encode("utf-8")


def parse_yaml_bytes(raw: bytes):
    try:
        return yaml.safe_load(raw.decode("utf-8", errors="replace"))
    except yaml.YAMLError as e:
        api_error("INVALID_YAML", f"Некорректный YAML: {e}", status=422)


def serialize_yaml_bytes(obj) -> bytes:
    return yaml.safe_dump(obj, allow_unicode=True, sort_keys=False).encode("utf-8")


def parse_xlsx_bytes(raw: bytes) -> list:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        api_error("INVALID_XLSX", f"Не удалось открыть XLSX: {e}", status=422)

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter, [])]
    if not headers:
        return []

    return [dict(zip(headers, row)) for row in rows_iter]


def serialize_xlsx_bytes(obj) -> bytes:
    headers, table = _row_dicts_to_table(obj)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in table:
        ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


DATA_PARSERS = {
    "csv": lambda raw: parse_delimited_bytes(raw, ","),
    "tsv": lambda raw: parse_delimited_bytes(raw, "\t"),
    "json": parse_json_bytes,
    "xml": parse_xml_bytes,
    "xlsx": parse_xlsx_bytes,
    "yaml": parse_yaml_bytes,
    "yml": parse_yaml_bytes,
}

DATA_SERIALIZERS = {
    "csv": lambda obj: serialize_delimited_bytes(obj, ","),
    "tsv": lambda obj: serialize_delimited_bytes(obj, "\t"),
    "json": serialize_json_bytes,
    "xml": serialize_xml_bytes,
    "xlsx": serialize_xlsx_bytes,
    "yaml": serialize_yaml_bytes,
    "yml": serialize_yaml_bytes,
}


def convert_data_bytes(raw: bytes, src_ext: str, target: str) -> bytes:
    obj = DATA_PARSERS[src_ext](raw)
    return DATA_SERIALIZERS[target](obj)


# ---------------------------------------------------------
# Document conversions
# ---------------------------------------------------------
def docx_to_pdf_via_libreoffice(docx_bytes: bytes) -> bytes:
    # Requires `soffice` available in PATH (installed by backend/Dockerfile).
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        in_path = tmp_path / "input.docx"
        out_dir = tmp_path / "out"
        out_dir.mkdir(parents=True, exist_ok=True)

        in_path.write_bytes(docx_bytes)

        cmd = [
            "soffice",
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            "--convert-to",
            "pdf",
            "--outdir",
            str(out_dir),
            str(in_path),
        ]
        try:
            subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=LIBREOFFICE_TIMEOUT_SECONDS,
            )
        except FileNotFoundError:
            api_error(
                "LIBREOFFICE_NOT_FOUND",
                "Не найден soffice (LibreOffice).",
                status=500,
            )
        except subprocess.TimeoutExpired:
            api_error(
                "DOCX_TO_PDF_TIMEOUT",
                "Конвертация DOCX→PDF превысила допустимое время.",
                status=504,
            )
        except subprocess.CalledProcessError as e:
            api_error(
                "DOCX_TO_PDF_FAILED",
                f"Ошибка конвертации DOCX→PDF: {e.stderr.decode('utf-8', errors='replace')[:500]}",
                status=500,
            )

        pdf_path = out_dir / "input.pdf"
        if not pdf_path.exists():
            api_error("DOCX_TO_PDF_FAILED", "LibreOffice не создал PDF файл.", status=500)

        return pdf_path.read_bytes()


def pdf_to_docx_via_pdf2docx(pdf_bytes: bytes) -> bytes:
    # Best effort conversion
    from pdf2docx import Converter

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        in_path = tmp_path / "input.pdf"
        out_path = tmp_path / "output.docx"
        in_path.write_bytes(pdf_bytes)

        try:
            cv = Converter(str(in_path))
            cv.convert(str(out_path), start=0, end=None)
            cv.close()
        except Exception as e:
            api_error("PDF_TO_DOCX_FAILED", f"Ошибка конвертации PDF→DOCX: {str(e)[:400]}", status=500)

        if not out_path.exists():
            api_error("PDF_TO_DOCX_FAILED", "Конвертер не создал DOCX файл.", status=500)

        return out_path.read_bytes()


# ---------------------------------------------------------
# POST /convert/media  (batch up to 20)
# ---------------------------------------------------------
@router.post("/media")
async def convert_media(
        files: list[UploadFile] = File(...),
        target: Literal["png", "jpeg", "jpg", "webp", "gif", "bmp", "tiff", "avif"] = Form(...)
):
    if not files:
        api_error("NO_FILES", "Не переданы файлы.", field="files", status=422)

    if len(files) > 20:
        api_error("TOO_MANY_FILES", "Для медиа можно загрузить максимум 20 файлов.", field="files", status=422)

    target_ext = "jpg" if target == "jpeg" else target
    outputs: list[tuple[str, bytes]] = []
    total_input = 0
    total_output = 0

    try:
        for f in files:
            ext = safe_ext(f.filename or "")
            if ext not in IMAGE_INPUT_EXTS:
                api_error(
                    "UNSUPPORTED_INPUT",
                    f"Неподдерживаемый тип файла для media: .{ext}",
                    field="files",
                    status=422,
                )

            raw = await read_upload_limited(f, MEDIA_MAX_FILE_SIZE, field="files")
            total_input += len(raw)
            if total_input > MEDIA_MAX_TOTAL_INPUT_SIZE:
                api_error(
                    "REQUEST_TOO_LARGE",
                    f"Суммарный размер media-файлов не должен превышать {MEDIA_MAX_TOTAL_INPUT_SIZE} байт.",
                    field="files",
                    status=413,
                )

            try:
                async with _conversion_slots:
                    out_bytes = await run_in_threadpool(convert_image_bytes, raw, ext, target_ext)
            except Exception as e:
                from fastapi import HTTPException
                if isinstance(e, HTTPException):
                    raise
                api_error("CONVERT_FAILED", f"Ошибка конвертации изображения: {str(e)[:300]}", status=500)

            total_output += len(out_bytes)
            if total_output > MEDIA_MAX_TOTAL_OUTPUT_SIZE:
                api_error(
                    "OUTPUT_TOO_LARGE",
                    f"Суммарный размер результата не должен превышать {MEDIA_MAX_TOTAL_OUTPUT_SIZE} байт.",
                    status=413,
                )

            out_name = f"{base_name(f.filename or 'file')}.{target_ext}"
            outputs.append((out_name, out_bytes))
    finally:
        for f in files:
            await close_upload(f)

    if len(outputs) == 1:
        name, data = outputs[0]
        return streaming_download(data, name, guess_mime_for_ext(target_ext))

    zip_data = await run_in_threadpool(zip_bytes, outputs)
    return streaming_download(zip_data, "converted_media.zip", "application/zip")


# ---------------------------------------------------------
# POST /convert/data  (single file)
# ---------------------------------------------------------
@router.post("/data")
async def convert_data(
        file: UploadFile = File(...),
        target: Literal["csv", "tsv", "json", "xml", "xlsx", "yaml", "yml"] = Form(...)
):
    if not file:
        api_error("NO_FILE", "Не передан файл.", field="file", status=422)

    src_ext = safe_ext(file.filename or "")
    if src_ext not in DATA_INPUT_EXTS:
        api_error("UNSUPPORTED_INPUT", f"Неподдерживаемый входной формат: .{src_ext}", field="file", status=422)

    if target not in DATA_TARGET_EXTS:
        api_error("UNSUPPORTED_TARGET", f"Неподдерживаемый целевой формат: {target}", field="target", status=422)

    try:
        raw = await read_upload_limited(file, DATA_MAX_FILE_SIZE)
    finally:
        await close_upload(file)

    # csv->csv, json->json etc. still round-trip through parse+serialize so
    # malformed same-format input is rejected instead of returned unchanged.
    async with _conversion_slots:
        out = await run_in_threadpool(convert_data_bytes, raw, src_ext, target)

    out_name = f"{base_name(file.filename or 'file')}.{target}"
    return streaming_download(out, out_name, guess_mime_for_ext(target))


# ---------------------------------------------------------
# POST /convert/document  (single file, doc/pdf)
# ---------------------------------------------------------
@router.post("/document")
async def convert_document(
        file: UploadFile = File(...),
        target: Literal["docx", "pdf"] = Form(...)
):
    if not file:
        api_error("NO_FILE", "Не передан файл.", field="file", status=422)

    src_ext = safe_ext(file.filename or "")
    if src_ext not in DOC_INPUT_EXTS:
        api_error("UNSUPPORTED_INPUT", f"Неподдерживаемый входной формат: .{src_ext}", field="file", status=422)

    if target not in DOC_TARGET_EXTS:
        api_error("UNSUPPORTED_TARGET", f"Неподдерживаемый целевой формат: {target}", field="target", status=422)

    if src_ext == target:
        api_error("NOOP", "Входной формат уже совпадает с целевым.", status=422)

    try:
        raw = await read_upload_limited(file, DOC_MAX_FILE_SIZE)
    finally:
        await close_upload(file)

    async with _conversion_slots:
        if src_ext == "docx" and target == "pdf":
            out = await run_in_threadpool(docx_to_pdf_via_libreoffice, raw)
        elif src_ext == "pdf" and target == "docx":
            out = await run_in_threadpool(pdf_to_docx_via_pdf2docx, raw)
        else:
            api_error("UNSUPPORTED_CONVERSION", f"Конвертация {src_ext} → {target} не поддержана.", status=422)

    if len(out) > DOC_MAX_OUTPUT_SIZE:
        api_error(
            "OUTPUT_TOO_LARGE",
            f"Размер результата не должен превышать {DOC_MAX_OUTPUT_SIZE} байт.",
            status=413,
        )

    out_name = f"{base_name(file.filename or 'file')}.{target}"
    return streaming_download(out, out_name, guess_mime_for_ext(target))
